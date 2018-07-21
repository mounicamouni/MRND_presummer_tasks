import openpyxl

wb = openpyxl.load_workbook('students.xlsx')
#print(type(wb))
#print(wb.get_sheet_names())
l=wb.get_sheet_names()
l.pop()
#anotherSheet = wb.active
#print(anotherSheet)
#sheet = wb.get_sheet_by_name('Current')
#print(sheet)
#print(sheet['A1'].value)

#print(sheet.max_row)
#print(sheet.max_column)

#print(sheet.cell(row=1, column=2).value)

wb1 = openpyxl.Workbook()
print(wb1.active)

sheet1 = wb1.get_sheet_by_name('Sheet')
wb1.remove_sheet(sheet1)
for i in l:
    sheet = wb.get_sheet_by_name(i)
    sheet1=wb1.create_sheet(title=i)
    n = sheet.max_row
    print(n)
    m = sheet.max_column
    print(m)
    for i in range(1,n+1):
        for j in range(1,m+1):
            sheet1.cell(row=i, column=j).value=sheet.cell(row=i, column=j).value.upper()
wb1.save("studentsCopy2.xlsx")