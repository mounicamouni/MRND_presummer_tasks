import openpyxl
from openpyxl.styles import Font
from copy import copy
import click
import os

@click.command()
@click.option('--capitalize/--no-capitalize',default=False,help="Capitalize the contents")
@click.option('--preservestyles/--no-preservestyles',default=False,help="Preserve the style")
@click.argument('src',nargs=1)
@click.argument('des',nargs=1)
def copyexcel(capitalize,preservestyles,src,des):
    wb = openpyxl.load_workbook(src)

    l=wb.get_sheet_names()
    l.pop()
    wb1 = openpyxl.Workbook()
    print(wb1.active)

    sheet1 = wb1.get_sheet_by_name('Sheet')
    wb1.remove_sheet(sheet1)
    for i in l:
        sheet = wb.get_sheet_by_name(i)
        sheet1=wb1.create_sheet(title=i)
        n = sheet.max_row
        m = sheet.max_column
        if capitalize:
            for i in range(1,n+1):
                for j in range(1,m+1):
                    sheet1.cell(row=i, column=j).value=sheet.cell(row=i, column=j).value.capitalize()
        else:
            for i in range(1,n+1):
                for j in range(1,m+1):
                    sheet1.cell(row=i, column=j).value=sheet.cell(row=i, column=j).value

        if preservestyles:
            for i in range(1,n+1):
                for j in range(1,m+1):
                    sheet1.cell(row=i, column=j).font=copy((sheet.cell(row=i, column=j)).font)
                    sheet1.cell(row=i, column=j).border = copy(sheet.cell(row=i, column=j).border)
                    sheet1.cell(row=i, column=j).fill = copy(sheet.cell(row=i, column=j).fill)
                    sheet1.cell(row=i, column=j).number_format = copy(sheet.cell(row=i, column=j).number_format)
                    sheet1.cell(row=i, column=j).protection  = copy(sheet.cell(row=i, column=j).protection )
                    #sheet1.cell(row=i, column=j).alignment  = copy(sheet.cell(row=i, column=j).alignment )

                    sheet1.row_dimensions[i].height = copy(sheet.row_dimensions[i].height)
                    #sheet1.row_dimensions[i].width = copy(sheet.row_dimensions[i].width)
                    #sheet1.cell(row=i, column=j).width = copy(sheet.cell(row=i, column=j).width)
                    #sheet1.row_dimensions[i].height=30
                    #sheet1.column_dimensions['A'].width=82
                    #sheet1.column_dimensions=copy(sheet.column_dimensions)


    if(os.path.exists(des)):
        value = click.prompt('Do u want to replace file', type=int)
        print(value)
        if(value==1):
            wb1.save(des)
    else:
        wb1.save(des)

if __name__ == '__main__':
    copyexcel()