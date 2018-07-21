from bs4 import BeautifulSoup
import openpyxl
import click

@click.command()
@click.argument('src',nargs=1)
@click.argument('des',nargs=1)
def htmlcopy(src,des):

    ####python appsAss3.py mockresult.html soup.xlsx

    wb1 = openpyxl.Workbook()
    sheet1 = wb1.get_sheet_by_name('Sheet')
    wb1.remove_sheet(sheet1)
    sheet1=wb1.create_sheet(title='mockresult')

    soup = BeautifulSoup(open(src))

    i=1
    j=1

    x=soup.find_all('th')
    for data in x[1:]:
        sheet1.cell(row=i, column=j).value = data.get_text()
        j+=1
    i+=1

    #print(x)
    for t in soup.find_all('tr')[1:]:
        cols = t.find_all('td')
        cols = [ele.text.strip() for ele in cols]

        cols.pop(0)
        j=1
        sheet1.cell(row=i, column=j).value=cols[0]
        j+=1
        for data in cols[1:]:
            sheet1.cell(row=i, column=j).value = int(data)
            j += 1
        i += 1
        #print(cols)


    wb1.save(des)

if __name__=='__main__':
    htmlcopy()