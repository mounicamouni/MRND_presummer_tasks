import MySQLdb
import openpyxl
import click
def drop():
    cur.execute("drop database mrnd")

def create():
    cur.execute("create database mrnd")

    cur.execute('use mrnd')

    marks = "create table marks(dbname char(100), transform int(5), custom int(5), piglatin int(5),topchars int(5),total int(5),primary key(dbname)  )"
    cur.execute(marks)

    student="create table students(name char(100),college char(100), email char(100),dbname char(100),foreign key(dbname) references marks(dbname))"
    cur.execute(student)



##for importing...................
def importing():
    print("hyrtht")
    cur.execute('use mrnd')

    wb1 = openpyxl.load_workbook('soup.xlsx')
    sheet1 = wb1.get_sheet_by_name('mockresult')
    n = sheet1.max_row


    for i in range(2, n + 1):
        db=str(sheet1.cell(row=i, column=1).value)
        tr=int(sheet1.cell(row=i, column=2).value)
        cu=int(sheet1.cell(row=i, column=3).value)
        pi = int(sheet1.cell(row=i, column=4).value)
        top = int(sheet1.cell(row=i, column=5).value)
        tot = int(sheet1.cell(row=i, column=6).value)
        cur.execute("insert into  marks(dbname,transform,custom,piglatin,topchars,total) values(%s,%s,%s,%s,%s,%s)", (db,tr,cu,pi,top,tot))

    wb = openpyxl.load_workbook('students.xlsx')
    sheet = wb.get_sheet_by_name('Current')

    n = sheet.max_row

    for i in range(2, n + 1):
        n = sheet.cell(row=i, column=1).value
        c = sheet.cell(row=i, column=2).value
        e = sheet.cell(row=i, column=3).value
        d = "ol2016_" + c + "_" + sheet.cell(row=i, column=4).value.lower() + "_mock"
        cur.execute("insert into students(name,college,email,dbname) values(%s,%s,%s,%s)", (n, c, e, d))

    cur.close()
    conn.commit()
    conn.close()
def stats():
    cur.execute('use mrnd')
    query="select s.college,count(*),min(m.total),max(m.total),avg(m.total) from students s inner join marks m on m.dbname=s.dbname group by(college) "
    cur.execute(query)
    for i in cur.fetchall():
        print(i)
@click.command()
@click.option("--createdb/--no",default=False)
@click.option("--dropdb/--no",default=False)
@click.option("--importdata/--no",default=False)
@click.option("--collegestats/--no",default=False)
def sql(createdb,dropdb,importdata,collegestats):
    if(createdb):
        create()
    if(dropdb):
        drop()
    if(importdata):
        importing()
    if(collegestats):
        stats()


if __name__ == '__main__':
    conn = MySQLdb.connect('127.0.0.1', 'root', 'root', '')
    cur = conn.cursor()

    sql()

    cur.close()
    conn.commit()
    conn.close()

